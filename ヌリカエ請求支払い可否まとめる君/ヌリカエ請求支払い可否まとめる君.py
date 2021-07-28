####################ヌリカエはサイトの改修を意外と頻繁にやってくるので後々、このプログラムも弄る必要が出てくるかも####################

# #今回使用するライブラリはこちら
import re
import os
import csv
import glob
import time
import openpyxl
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
import chromedriver_binary
from selenium.webdriver.common.keys import Keys
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
import win32com.client


#ブラウザ制御のためにここでドライバ読み込む
driver = webdriver.Chrome()
# driver = webdriver.PhantomJS(executable_path='C:/Users/taked/phantomjs-2.1.1-windows/bin/phantomjs.exe')      #chromeだとGUIがいちいち立ち上がるので重い。こっちのドライバでいけばGUIが立ち上がらずに済む
#　↑※phantomjsはメンテナーが引退したので使用不可。別のを用意する


#目的のサイトに行き、ログイン認証を済ませる
driver.get('')                                           #行きたいサイト、今回はヌリカエのウェブサイトのURLを指定
driver.find_element_by_id('client_tool_account_identifier').send_keys('')       #フォームを取得してID入力
driver.find_element_by_id('client_tool_account_password').send_keys('')         #フォームを取得してパス入力　←これセキュリティ的にやべーね
driver.find_element_by_name('commit').click()                                           #フォームを送信してログイン


#請求明細のページを見に行く
driver.find_element_by_link_text('請求明細').click()
#印刷ページを開く
# driver.find_element_by_tag_name('form').click()


#BeautifulSoupで請求明細の必要な情報を抜き出す
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')
td_data = soup.select('tbody td')


#Excelファイルを新規作成
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'test'
ws1 = wb.worksheets[0]


# BeautifulSoupで取得しておいた請求明細のtbodyタグのtd要素（すなわち紹介料と成約料）をエクセルにコピー。不要なので成約料は削除
row_count = 4
for i in range(4, len(td_data) - 4, 4):
    ws1.cell(row = row_count, column = 2, value = td_data[i].text.strip())
    ws1.cell(row = row_count, column = 3, value = td_data[i + 1].text.strip())
    ws1.cell(row = row_count, column = 4, value = td_data[i + 2].text.strip())
    ws1.cell(row = row_count, column = 5, value = td_data[i + 3].text.strip())
    if '/' in ws1.cell(row = row_count, column = 2).value  and 'ご紹介料' in ws1.cell(row = row_count, column = 3).value:
        row_count += 1
    elif '' in ws1.cell(row = row_count, column = 2).value and '課金除外' in ws1.cell(row = row_count, column = 3).value:
        row_count += 1
    else:
        ws1.delete_rows(row_count)
        row_count += 1


#都道府県、市町村、支払いの可否の３つ分、列を追加
ws1.insert_cols(4,3)


#罫線の準備
border = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))

border2 = Border(top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000'),
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'))


#B1から一番下までの範囲を中央ぞろえした
for cell1 in ws1['B1:H{}'.format(ws1.max_row)]:
    for cell2 in cell1:
        cell2.alignment = Alignment(horizontal='center', vertical = 'center')
        cell2.font = Font(name = 'メイリオ', size = 14, bold = True)


#B4から一番下までの範囲を中央ぞろえした
for cell3 in ws1['B4:H{}'.format(ws1.max_row)]:
    for cell4 in cell3:
        cell4.border = border


#エクセルファイル　列や行の幅、フォントサイズなどを調整
ws1.row_dimensions[1].height = 36
ws1.row_dimensions[3].height = 24
ws1['D1'] = 'ヌリカエ今月分の請求です。ご確認宜しくお願いします'
ws1['D1'].font = Font(size = 20)
ws1['D1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 16
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 12

ws1['B3'] = '紹介日'
ws1['C3'] = '項目'
ws1['D3'] = '都道府県'
ws1['E3'] = '市区町村など'
ws1['F3'] = '支払いの可否'
ws1['G3'] = '税区分'
ws1['H3'] = '単価（税抜）'
ws1['I3'] = 'ブロックコード'


for font1 in ws1['B3:H3']:
    for font2 in font1:
        font2.font = Font(size = 10)
        font2.alignment = Alignment(horizontal = 'center', vertical = 'center')
        font2.border = border2


############################csvをサイトからとってきて開く################################################################################################################
#請求書の日付を手に入れる。以下は関数化できそう。
closing_day1 = soup.select('[class="billing_closing_day"] td') #請求書の締め日の日付をタグごと取得
closing_day2 = closing_day1[0].text.strip()  # <- 2020年8月20日   タグを削除
dt = datetime.strptime(closing_day2, '%Y年%m月%d日')   #文字列から年月日を分割して抽出  これ以降年月日の数値をdt.year dt.month dt.dayとして個別に扱える


driver.find_element_by_link_text('紹介一覧').click()    #紹介一覧のページへ移動

driver.find_element_by_class_name('panel-title').click()    #「条件で絞り込む」のパネルを開く
time.sleep(0.1)   #パネルが開くまで明示的に待たないと怒られる仕様なので待つ

#例えばclosing_dayが8月だったら7/21~8/20、9月だったら8/21~9/20を入力するようにしてる
start_date = driver.find_element_by_id('q_created_at_gteq')
if dt.month == 1:
    start_date.send_keys(dt.year - 1)
    start_date.send_keys(Keys.RIGHT)
else:
    start_date.send_keys(dt.year)
    start_date.send_keys(Keys.RIGHT)
if dt.month == 1:   #先月を参照するこの部分、今月が「1月」である場合の対策。抜かりないぜ。
    start_date.send_keys('12')
else:
    start_date.send_keys(dt.month - 1)
start_date.send_keys(Keys.RIGHT)
start_date.send_keys(dt.day + 1)

end_date = driver.find_element_by_id('q_created_at_lteq_end_of_day')
end_date.send_keys(dt.year)
end_date.send_keys(Keys.RIGHT)
end_date.send_keys(dt.month)
end_date.send_keys(Keys.RIGHT)
end_date.send_keys(dt.day)

driver.find_element_by_id('csv-output-btn').find_element_by_class_name('btn').click()
time.sleep(3)

###########################作成したExcelファイル(1)ととってきたcsvを変換？したExcelファイル(2)の内容を比較。名前をキーにして(2)から住所を抽出、(1)の対応する名前の横に追加#########

csv_list = glob.glob('')  #ダウンロードcsvを探してる  ここはもうちょっと上手いやり方を考える余地あり
target_csv = csv_list[0]

wb2 = openpyxl.Workbook()
ws2 = wb2.active

with open('{}'.format(target_csv)) as f:    #この辺でcsvの中身をExcelに移動させる
    reader = csv.reader(f)
    for row in reader:
        ws2.append(row)


#比較時に邪魔になるので「様 ご紹介料」の文字をws1の表から削除
for remove_gosyokairyo_row in range(ws1.max_row - (ws1.max_row - 4), ws1.max_row + 1):
    if ws1.cell(row = remove_gosyokairyo_row, column = 3).value:
        ws1.cell(row = remove_gosyokairyo_row, column = 3).value = ws1.cell(row = remove_gosyokairyo_row, column = 3).value.replace('様 ご紹介料', '').strip()
    else:
        pass


# 2つのエクセルファイルを比較して、名前が一致すれば住所をws2からws1へコピー
for hikaku_parent in range(ws1.max_row - (ws1.max_row - 4), ws1.max_row + 1):
    for hikaku_child in range(ws2.max_row - (ws2.max_row - 2), ws2.max_row + 1):
        if ws1.cell(row = hikaku_parent, column = 3).value.strip() == ws2.cell(row = hikaku_child, column = 2).value.strip():
            address = ws2.cell(row = hikaku_child, column = 3).value.strip()
            pref_city = re.match('(...??[都道府県])((?:旭川|伊達|石狩|盛岡|奥州|田村|南相馬|那須塩原|東村山|武蔵村山|羽村|十日町| \
                                    上越|富山|野々市|大町|蒲郡|四日市|姫路|大和郡山|廿日市|下松|岩国|田川|大村)市|.+?郡(?:玉村|大町| \
                                    .+?)[町村]|.+?市.+?区|.+?[市区町村])(.*)', address)
            try:
                ws1.cell(row = hikaku_parent, column = 4).value = pref_city.groups()[0]
                ws1.cell(row = hikaku_parent, column = 5).value = pref_city.groups()[1]
            except AttributeError:
                ws1.cell(row = hikaku_parent, column = 4).value = "住所"
                ws1.cell(row = hikaku_parent, column = 5).value = "マッチせず"


#ブロックコードをつくる
hokkaido = {'北海道'}
tohoku1 = {'青森県', '岩手県', '秋田県'}
tohoku2 = {'宮城県', '山形県', '福島県', '新潟県'}
kantou1 = {'茨城県', '栃木県', '埼玉県', '群馬県', '富山県', '石川県'}
tokyo1 = {'青梅市','八王子市','武蔵村山市','西多摩郡','昭島市','東大和市','立川市', \
          '羽村市','板橋区','福生市','あきる野市'}
tokyo = {'東京都'}
tokyo2 = {'練馬区','三鷹市','杉並区','東久留米市','国立市','小金井市','町田市','世田谷区', \
          '狛江市','西東京市','葛飾区','小平市','調布市','府中市','中野区','稲城市', \
          '江戸川区','墨田区','文京区','国分寺市','大田区','品川区','新宿区'}
kantou2 = {'千葉県', '神奈川県', '静岡県', '山梨県'}
niigata_for_nagano = {'上越市', '妙高市'}
tyubu = {'愛知県', '長野県', '岐阜県'}
mie = {'三重県'}
kinki = {'滋賀県', '京都府', '奈良県', '和歌山県', '福井県'}
osaka = {'大阪府'}
hyogo_for_osaka = {'丹波市','西宮市','芦屋市','伊丹市','三田市','篠山市','宝塚市','川西市'}
hyogo_osaka_dottika = {'西脇市', '大阪市'}
shikoku_sanin = {'兵庫県', '鳥取県', '島根県', '岡山県', '広島県', '徳島県', '香川県', '愛媛県', '高知県'}
kyusyu = {'山口県', '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県'}

for block1 in range(4, ws2.max_row):    #if文は順番大事
    if ws1.cell(row = block1, column = 4).value in hokkaido:
        ws1.cell(row = block1, column = 9).value = '01'
    elif ws1.cell(row = block1, column = 5).value in niigata_for_nagano:
        ws1.cell(row = block1, column = 9).value = '09'
    elif ws1.cell(row = block1, column = 5).value in hyogo_for_osaka:
        ws1.cell(row = block1, column = 9).value = '14'
    elif ws1.cell(row = block1, column = 5).value in hyogo_osaka_dottika:
        ws1.cell(row = block1, column = 9).value = '15'
    elif ws1.cell(row = block1, column = 4).value in tohoku1:
        ws1.cell(row = block1, column = 9).value = '02'
    elif ws1.cell(row = block1, column = 4).value in tohoku2:
        ws1.cell(row = block1, column = 9).value = '03'
    elif ws1.cell(row = block1, column = 4).value in kantou1:
        ws1.cell(row = block1, column = 9).value = '04'
    elif ws1.cell(row = block1, column = 5).value in tokyo1:
        ws1.cell(row = block1, column = 9).value = '05'
    elif ws1.cell(row = block1, column = 5).value in tokyo2:
        ws1.cell(row = block1, column = 9).value = '07'
    elif ws1.cell(row = block1, column = 4).value in tokyo:
        ws1.cell(row = block1, column = 9).value = '06'
    elif ws1.cell(row = block1, column = 4).value in kantou2:
        ws1.cell(row = block1, column = 9).value = '08'
    elif ws1.cell(row = block1, column = 4).value in tyubu:
        ws1.cell(row = block1, column = 9).value = '10'
    elif ws1.cell(row = block1, column = 4).value in mie:
        ws1.cell(row = block1, column = 9).value = '11'
    elif ws1.cell(row = block1, column = 4).value in kinki:
        ws1.cell(row = block1, column = 9).value = '12'
    elif ws1.cell(row = block1, column = 4).value in osaka:
        ws1.cell(row = block1, column = 9).value = '13'
    elif ws1.cell(row = block1, column = 4).value in shikoku_sanin:
        ws1.cell(row = block1, column = 9).value = '16'
    elif ws1.cell(row = block1, column = 4).value in kyusyu:
        ws1.cell(row = block1, column = 9).value = '17'
    else:
        ws1.cell(row = block1, column = 9).value = '18'


# タイトル行
header_cells = ws1[3]

# 4行目以降（データ） #紹介日や顧客名などの情報をsortできるようにリスト化する
nurikae_list = []
for row in ws1.iter_rows(min_row=4):
    if row[4].value is None:    #E列（適当に選んだ。row[1]でもrow[3]でもいいと思う）が空欄である行はスキップ
        continue
    row_dic = {}
    # セルの値を「key-value」で登録
    for k, v in zip(header_cells, row):
        row_dic[k.value] = v.value
    nurikae_list.append(row_dic)

nurikae_list_sorted = sorted(nurikae_list, key=lambda x: x['ブロックコード'])

#リストの中身が辞書なので項目がバラバラになる、したがってキーで検索して並べていく必要がある
for list_count in range(4, len(nurikae_list_sorted) + 4):
    ws1.cell(row = list_count, column = 2).value = nurikae_list_sorted[list_count - 4]['紹介日']
    ws1.cell(row = list_count, column = 3).value = nurikae_list_sorted[list_count - 4]['項目']
    ws1.cell(row = list_count, column = 4).value = nurikae_list_sorted[list_count - 4]['都道府県']
    ws1.cell(row = list_count, column = 5).value = nurikae_list_sorted[list_count - 4]['市区町村など']
    ws1.cell(row = list_count, column = 7).value = nurikae_list_sorted[list_count - 4]['税区分']
    ws1.cell(row = list_count, column = 8).value = nurikae_list_sorted[list_count - 4]['単価（税抜）']


#不要になるのでブロックコードの列削除
ws1.delete_cols(9)


#印刷設定   pdfに変換するときにも影響する
# ws1.print_title_rows = '1:1'              # 印刷タイトル行
ws1.page_setup.orientation = 'portrait'  # 横は'landscape'
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.sheet_properties.pageSetUpPr.fitToPage = True


#必要なやつだけセーブ
wb.save('test.xlsx')    #このpythonファイルをダブルクリックで実行するならこのファイルパス　vscode上で実行するなら  'ヌリカエ請求支払い可否まとめる君/test.xlsx'　とする。でも絶対パスにするのが一番わかりやすいかもね


#要らないのでダウンロードしたcsv削除
# os.remove(target_csv)


# ブラウザを閉じる
driver.close()


# #Excelファイルをpdfに変換
# excel = win32com.client.Dispatch("Excel.Application")
# file = excel.Workbooks.Open("C:/Users/taked/Desktop/python/ヌリカエ請求支払い可否まとめる君/test.xlsx")
# file.WorkSheets(1).Select()
# file.ActiveSheet.ExportAsFixedFormat(0,"C:/Users/taked/Desktop/python/ヌリカエ請求支払い可否まとめる君/test.pdf")
# file.close()