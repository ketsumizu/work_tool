##あえてウザいぐらいコメントを書いてみる


# #モジュールをインポート
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.workbook import Workbook

# #ファイルを開く   　　　　　　　　　　　　　　　　　　ファイル名をどうするかは依頼者らと要相談
wb = openpyxl.load_workbook('.xlsx')

# #１番目のシートを選択
ws1 = wb.worksheets[0]

#先頭列の色を白にする
for row in ws1['A1':'AC1']:
    for cell in row:
        cell.fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFFFFF',bgColor='FFFFFF')

# #１番目のシートを右端の位置へとコピー
wb.copy_worksheet(ws1)

# #２番目のシートを選択
ws2 = wb.worksheets[1]

# #不要な列を削除
ws2.delete_cols(1,5)
ws2.delete_cols(13,1)

# #1行目に行を挿入×2回
ws2.insert_rows(1)
ws2.insert_rows(1)

#２つ目のシート、上3つのセルの高さをいい感じに調整
ws2.row_dimensions[1].height = 21.75
ws2.row_dimensions[2].height = 21.75
ws2.row_dimensions[3].height = 60

#２つ目のシート、セルに文字を書き込む
ws2['A1'] = 'ご利用明細'
ws2['B1'] = '2021年'
ws2['C1'] = '月分'
ws2['G1'] = ''
ws2['H1'] = '支店名'
ws2['I1'] = '※納品書と照合後、押印したものを本社2Fにファックスを入れる。'
ws2['I2'] = '※添付書類・・・ｵｲﾙ交換、修理・部品交換等本社承認済みの依頼書。灯油は納品書。'
ws2['V1'] = '照合者'
ws2['V2'] = '印'

#２つ目のシート、セルの行の幅をいい感じに調整
ws2.column_dimensions['A'].width = 7
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 3
ws2.column_dimensions['D'].width = 5
ws2.column_dimensions['E'].width = 4
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 15.5
ws2.column_dimensions['H'].width = 11
ws2.column_dimensions['I'].width = 3
ws2.column_dimensions['J'].width = 3
ws2.column_dimensions['K'].width = 3
ws2.column_dimensions['L'].width = 4
ws2.column_dimensions['M'].width = 20
ws2.column_dimensions['N'].width = 9
ws2.column_dimensions['O'].width = 9
ws2.column_dimensions['P'].width = 9
ws2.column_dimensions['Q'].width = 7
ws2.column_dimensions['R'].width = 7
ws2.column_dimensions['S'].width = 6
ws2.column_dimensions['T'].width = 7
ws2.column_dimensions['U'].width = 16
ws2.column_dimensions['V'].width = 9

#２つ目のシート、印の文字を中央ぞろえ
ws2['V1'].alignment = Alignment(horizontal='center')
ws2['V2'].alignment = Alignment(horizontal='center')

#ループさせるときに何回か使うから、この辺で適当に最終行が何行目か取得しとく
max = ws2.max_row

# 「届先小計」の文字列に完全一致した数をカウントしてT1のセルに入れる。
glo = 0   #うわ！グローバル変数使ってもうた！リファクタリングせな！
for todo_count in range(1, max):
    todo_count2 = ws2.cell(row=todo_count, column=14).value
    if todo_count2 == "届先小計":
        glo += 1

#この後、データをコピーする下準備として、届先小計と同じ数のシートが必要になるので、それらを作成して、不要なデータを削除する。
for overwrite in range(glo):      #上で宣言したグローバル変数 k ここでも使ってもうた。後戻りできひん。あとこの辺工夫したらプログラムの速度もっと改善できそう。とりあえず動いてるからヨシ！
    wb.copy_worksheet(ws2)
    m = overwrite + 2
    wsx = wb.worksheets[m]
    for row in wsx['A4:W500']:
        for cell in row:
            cell.value = ""

# forループで、'届先小計'と完全一致したセルを見つけたらその位置から上、支店ごとのデータ範囲をコピー。それぞれの専用シートへ貼り付ける

todo_sum, worksheet_number, previous_row = 0, 2, 0    #届先小計カウント用変数とワークシートカウント用変数と1つ前のcurrent_rowの行数を保持しておく入れ物をアンパックを用いて宣言してみた

for x in range(0, max):
    if not worksheet_number > (glo + 1):    #ワークシートが前段階で生成した数を超えてなければ処理へ進む
        target_worksheet = wb.worksheets[worksheet_number]
    target = ws2.cell(row=x+4, column=14).value    #14列目を4行目から下にかけて確認していく
    current_row = x+4
    if todo_sum == glo:
        break
    elif target == '届先小計' and todo_sum == 0:
        for i in range(1,24):                                                       #列方向の範囲
            for j in range(4, current_row + 1):                                     #行方向の範囲
                copy = ws2.cell(row = j, column =i).value                           #コピー
                target_worksheet.cell(row = j, column =i , value = copy)            #貼り付け
        previous_row = current_row                                                  #今見てるcurrent_rowをprevious_rowに放り込んで次へ進む
        todo_sum += 1                                                               #届先小計のカウントを増やす
        worksheet_number += 1                                                       #コピー先を次のワークシートへ切り替え
    elif target == '届先小計' and todo_sum != 0:                                     #届先小計を見つけたときの２回目以降の処理
        for i in range(1,24):                                                       #列方向の範囲
            for j in range(previous_row, current_row + 1):                          #行方向の範囲
                copy = ws2.cell(row = j + 1, column = i).value                      #コピー
                target_worksheet.cell(row = j, column = i , value = copy)           #貼り付け　
        target_worksheet.delete_rows(idx = 4, amount = previous_row - 3)            #不要な行を削除
        previous_row = current_row
        todo_sum += 1
        worksheet_number += 1

#シート名を対応する支店名へ変更
ws2.title = '原本'

for title_number in range(2, glo + 2):
    ws3 = wb.worksheets[title_number]
    ws4 = ws3['A4'].value
    ws3.title = ws4

#G1のセルにシート名（対応する支店名）を入れる
branch = wb.sheetnames

for branch_number in range(2, glo + 2):
    ws5 = wb.worksheets[branch_number]
    ws6 = branch[branch_number]
    ws5['G1'] = ws6

#書式設定されていると空欄でも最終行と判断されるためmax_rowだと300行くらいが取得される。なので、見た目でわかる最終行（値の入っている最終行）の行数を取得
waka_row = wb['和歌山支店'].max_row
while wb['和歌山支店'].cell(row = waka_row, column = 1).value is None or not str(wb['和歌山支店'].cell(row = waka_row, column = 1).value).strip():
    waka_row -= 1

for minami in range(2, glo + 2): #シートを見て回って和歌山支店と南大阪支店が両方存在する場合、南大阪支店のデータを和歌山支店のデータの最終行の２つ下の行へコピーする
    if wb.worksheets[minami].title == '南大阪支店' and wb['和歌山支店']:
        for i in range(1,23):
            for j in range(4, current_row + 1):
                copy = wb.worksheets[minami].cell(row = j, column =i).value
                wb['和歌山支店'].cell(row = waka_row + j - 1, column =i , value = copy)


#罫線の余りを削除
border = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))


for delete_line in range(2, glo + 2):
    ws_z = wb.worksheets[delete_line]
    ws_row = wb.worksheets[delete_line].max_row
    while wb.worksheets[delete_line].cell(row = ws_row, column = 1).value is None or not str(wb.worksheets[delete_line].cell(row = ws_row, column = 1).value).strip():
        ws_row -= 1
    target_ws = wb.worksheets[delete_line]
    target_ws.delete_rows(idx = ws_row + 3, amount = target_ws.max_row)
    for underbar1 in ws_z['A{}:W{}'.format(target_ws.max_row, target_ws.max_row)]:
        for underbar2 in underbar1:
            underbar2.border = border


#不要になるのでシート内の支店名を削除
for delete_branch in range(1, glo + 2):
    wb.worksheets[delete_branch].delete_cols(1)



# if wb['南大阪支店'] and wb['和歌山支店']:       #データを和歌山シートへ移した後に不要になった南大阪シートを削除
#     wb.remove_sheet(wb.get_sheet_by_name('南大阪支店'))

#ファイル内の全てのシートをループして検索
check = False
for worksheets in wb.worksheets:
    #指定シートが存在していれば、変数にTrueを格納
    if worksheets.title == '南大阪支店':
        check = True

if check == True:
    wb.remove_sheet(wb.get_sheet_by_name('南大阪支店'))
else:
    pass


# #ファイルを上書き保存
wb.save('.xlsx')